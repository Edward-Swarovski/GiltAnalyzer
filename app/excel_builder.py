from __future__ import annotations

from collections.abc import Iterable

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font
from openpyxl.worksheet.worksheet import Worksheet

from app import formulas
from app.models import GiltMarketRow
from app.tax_scenarios import DEFAULT_TAX_SCENARIOS

DEFAULT_NOMINAL_AMOUNT = 10_000

SETTINGS_HEADERS = ("Setting", "Value", "Notes")

MARKET_DATA_HEADERS = (
    "ISIN",
    "Gilt Name",
    "Maturity",
    "Coupon %",
    "Imported Price",
    "Imported Yield %",
    "Valuation Date",
    "Source",
)

INPUT_HEADERS = ("ISIN", "Gilt Name", "Nominal Amount (£)", "Override Price", "Override Yield %")

ANALYSIS_HEADERS = (
    "ISIN",
    "Gilt Name",
    "Maturity",
    "Coupon %",
    "Effective Price",
    "Effective Yield %",
    "Approx Retail Ask Yield %",
    "DMO Retail Sale Clean Price",
    "DMO Retail Sale Dirty Price",
    "Nominal Amount (£)",
    "Years to Maturity",
    "Annual Coupon Cash (£)",
    "Capital Uplift to Par (£)",
    "Approx Gross Cash Gain to Maturity (£)",
    "Coupon Tax @20% (£)",
    "Coupon Tax @40% (£)",
    "Coupon Tax @45% (£)",
    "Approx Net Cash Gain @20% (£)",
    "Approx Net Cash Gain @40% (£)",
    "Approx Net Cash Gain @45% (£)",
    "Annual Net @20% (£)",
    "Annual Net @40% (£)",
    "Annual Net @45% (£)",
    "Approx Accrued Interest (£)",
)


def build_workbook(
    rows: Iterable[GiltMarketRow],
    *,
    default_nominal_amount: int = DEFAULT_NOMINAL_AMOUNT,
    retail_ask_yields: dict[str, float] | None = None,
    retail_quotes: dict | None = None,
) -> Workbook:
    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    settings = workbook.create_sheet("Settings")
    market_data = workbook.create_sheet("Market Data")
    inputs = workbook.create_sheet("Inputs")
    analysis = workbook.create_sheet("Analysis")
    instructions = workbook.create_sheet("Instructions")

    _write_headers(settings, SETTINGS_HEADERS)
    _write_headers(market_data, MARKET_DATA_HEADERS)
    _write_headers(inputs, INPUT_HEADERS)
    _write_headers(analysis, ANALYSIS_HEADERS)

    materialized_rows = list(rows)
    resolved_yields = retail_ask_yields or {}
    resolved_quotes = retail_quotes or {}
    _populate_settings(settings, default_nominal_amount)
    _populate_market_data(market_data, materialized_rows)
    _populate_inputs(inputs, materialized_rows)
    _populate_analysis(analysis, materialized_rows, resolved_yields, resolved_quotes)

    _populate_instructions(instructions)

    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions

    _format_analysis(analysis, len(materialized_rows))

    return workbook


# Columns (1-based) whose values are £ amounts — formatted to 2 decimal places.
# H=8(clean), I=9(dirty) are prices (not £ cash), J=10..Q=17 cash, R=18..T=20 net, U=21..W=23 annual net
_GBP_COLUMNS = (12, 13, 14, 15, 16, 17, 18, 19, 20, 21, 22, 23, 24)
_GBP_FORMAT = '£#,##0.00'
_ANALYSIS_FONT_SIZE = 9


def _format_analysis(sheet: Worksheet, data_row_count: int) -> None:
    # Word-wrap header row
    for cell in sheet[1]:
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    _BLUE = "0070C0"
    _ANNUAL_NET_COLS = (21, 22, 23)  # U, V, W

    # Word-wrap + blue header for Annual Net columns
    for col in _ANNUAL_NET_COLS:
        sheet.cell(1, col).font = Font(bold=True, color=_BLUE)

    # Font size + £ format on data rows; blue for Annual Net columns
    for r in range(2, data_row_count + 2):
        for cell in sheet[r]:
            cell.font = Font(size=_ANALYSIS_FONT_SIZE)
        for col in _GBP_COLUMNS:
            sheet.cell(r, col).number_format = _GBP_FORMAT
        for col in _ANNUAL_NET_COLS:
            sheet.cell(r, col).font = Font(size=_ANALYSIS_FONT_SIZE, color=_BLUE)

    # Print: fit all columns onto 1 page wide, portrait, landscape if needed
    sheet.page_setup.fitToPage = True
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0  # unlimited pages tall
    sheet.sheet_properties.pageSetUpPr.fitToPage = True


_RETAIL_ASK_YIELD_NOTE = (
    "Pre-computed import from DMO D10B sale dirty price.\n"
    "This value does NOT update when price or yield overrides change.\n"
    "Day count: days/365.25 (approximation — not exact Actual/Actual ICMA)."
)


def _write_headers(sheet: Worksheet, headers: tuple[str, ...]) -> None:
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True)
    if "Approx Retail Ask Yield %" in headers:
        col_idx = headers.index("Approx Retail Ask Yield %") + 1
        col_letter = sheet.cell(1, col_idx).column_letter
        cell = sheet[f"{col_letter}1"]
        cell.comment = Comment(_RETAIL_ASK_YIELD_NOTE, "GiltAnalyzer")


def _populate_settings(sheet: Worksheet, default_nominal_amount: int) -> None:
    sheet.append(("DefaultNominalAmount", default_nominal_amount, "Default £ face value applied to all gilts unless overridden in the Inputs sheet (column B)."))
    sheet.append((None, None, None))
    sheet.append(("--- Tax Scenarios ---", None, None))
    for scenario in DEFAULT_TAX_SCENARIOS:
        sheet.append((scenario.name, float(scenario.coupon_tax_rate), scenario.notes))


def _populate_market_data(sheet: Worksheet, rows: list[GiltMarketRow]) -> None:
    for row in rows:
        sheet.append(
            (
                row.isin,
                row.gilt_name,
                row.maturity_date,
                float(row.coupon_rate),
                float(row.imported_price) if row.imported_price is not None else None,
                float(row.imported_yield) if row.imported_yield is not None else None,
                row.valuation_date,
                row.source_name,
            )
        )


def _populate_inputs(sheet: Worksheet, rows: list[GiltMarketRow]) -> None:
    for excel_row, row in enumerate(rows, start=2):
        sheet.cell(excel_row, 1, row.isin)
        # Gilt Name: read-only lookup from Market Data — do not edit
        sheet.cell(excel_row, 2, f"=IFERROR(INDEX('Market Data'!B:B,MATCH(A{excel_row},'Market Data'!A:A,0)),\"\")")
        # Cols C, D, E: user-editable (Nominal Amount, Override Price, Override Yield)


def _populate_analysis(
    sheet: Worksheet,
    rows: list[GiltMarketRow],
    retail_ask_yields: dict[str, float],
    retail_quotes: dict,
) -> None:
    for excel_row, row in enumerate(rows, start=2):
        quote = retail_quotes.get(row.isin)
        sheet.cell(excel_row, 1, row.isin)
        sheet.cell(excel_row, 2, row.gilt_name)
        sheet.cell(excel_row, 3, row.maturity_date)
        sheet.cell(excel_row, 4, float(row.coupon_rate))
        sheet.cell(excel_row, 5, formulas.effective_price(excel_row))
        sheet.cell(excel_row, 6, formulas.effective_yield(excel_row))
        sheet.cell(excel_row, 7, retail_ask_yields.get(row.isin))
        sheet.cell(excel_row, 8, float(quote.sale_clean_price) if quote else "N/A")
        sheet.cell(excel_row, 9, float(quote.sale_dirty_price) if quote else "N/A")
        sheet.cell(excel_row, 10, formulas.nominal_amount(excel_row))
        sheet.cell(excel_row, 11, formulas.years_to_maturity(excel_row))
        sheet.cell(excel_row, 12, formulas.annual_coupon_cash(excel_row))
        sheet.cell(excel_row, 13, formulas.capital_uplift_to_par(excel_row))
        sheet.cell(excel_row, 14, formulas.approx_gross_cash_gain_to_maturity(excel_row))
        sheet.cell(excel_row, 15, formulas.coupon_tax(excel_row, "20%"))
        sheet.cell(excel_row, 16, formulas.coupon_tax(excel_row, "40%"))
        sheet.cell(excel_row, 17, formulas.coupon_tax(excel_row, "45%"))
        sheet.cell(excel_row, 18, formulas.approx_net_cash_gain(excel_row, "O"))  # R = N - O
        sheet.cell(excel_row, 19, formulas.approx_net_cash_gain(excel_row, "P"))  # S = N - P
        sheet.cell(excel_row, 20, formulas.approx_net_cash_gain(excel_row, "Q"))  # T = N - Q
        sheet.cell(excel_row, 21, formulas.annual_net_gain(excel_row, "R"))        # U = R / K
        sheet.cell(excel_row, 22, formulas.annual_net_gain(excel_row, "S"))        # V = S / K
        sheet.cell(excel_row, 23, formulas.annual_net_gain(excel_row, "T"))        # W = T / K
        sheet.cell(excel_row, 24, formulas.accrued_interest(excel_row))            # X



_INSTRUCTIONS = (
    ("GiltAnalyzer — How to use this workbook", True),
    ("", False),
    ("--- INPUTS SHEET ---", True),
    ("", False),
    ("ISIN (column A)", True),
    ("Read-only. Identifies each gilt. Do not edit.", False),
    ("", False),
    ("Gilt Name (column B)", True),
    ("Read-only formula — populated automatically from Market Data. Do not edit.", False),
    ("", False),
    ("Nominal Amount — column C", True),
    ("The £ face value of your holding in this gilt.", False),
    ("Leave blank to use the default from Settings!B2 (DefaultNominalAmount).", False),
    ("Enter a number (e.g. 50000) to override for this gilt only.", False),
    ("All cash flow columns in Analysis (J–R) scale with this value.", False),
    ("", False),
    ("Override Price — column D", True),
    ("Optional. Enter a clean price to replace the DividendData market price.", False),
    ("When set, Effective Price (Analysis col E) uses this value instead of the live price.", False),
    ("Leave blank to use the live market price.", False),
    ("", False),
    ("Override Yield % — column E", True),
    ("Optional. Enter a yield (in %) to replace the DividendData market yield.", False),
    ("When set, Effective Yield (Analysis col F) uses this value directly.", False),
    ("Leave blank to use the live market yield.", False),
    ("", False),
    ("--- ANALYSIS SHEET ---", True),
    ("", False),
    ("Price and yield columns (E, F, G, H, I)", True),
    ("", False),
    ("Effective Price (col E)", True),
    ("The mid-market clean price from DividendData — what the gilt trades at in the", False),
    ("  secondary market between brokers and investors, excluding accrued interest.", False),
    ("  'Clean' means accrued interest is not included — this is the quoted price.", False),
    ("  This is a mid price: halfway between the dealer buy (bid) and sell (offer) price.", False),
    ("  Can be overridden in Inputs col D if you have a specific broker quote.", False),
    ("", False),
    ("DMO Retail Sale Clean Price (col H)", True),
    ("The clean price at which the DMO will sell you a gilt through their retail service.", False),
    ("  This is the price shown on the DMO website — but NOT what you actually pay.", False),
    ("  Typically slightly higher than Effective Price — the DMO adds a spread.", False),
    ("  Shows N/A if no D10B file was loaded at export time.", False),
    ("", False),
    ("DMO Retail Sale Dirty Price (col I)", True),
    ("The actual cash you hand over per £100 nominal when buying through the DMO service.", False),
    ("  Dirty Price = Clean Price + Accrued Interest since the last coupon payment.", False),
    ("  Accrued interest builds up daily and resets to zero on each coupon payment date.", False),
    ("  This is the true settlement amount — what leaves your bank account.", False),
    ("  The Approx Retail Ask Yield % (col G) is derived from this dirty price.", False),
    ("  Shows N/A if no D10B file was loaded at export time.", False),
    ("", False),
    ("How the three prices relate:", True),
    ("  Effective Price (E)        ← market mid clean price  [DividendData]", False),
    ("       ↓ DMO adds a spread", False),
    ("  DMO Sale Clean Price (H)   ← DMO retail clean price  [D10B, slightly higher]", False),
    ("       ↓ + accrued interest", False),
    ("  DMO Sale Dirty Price (I)   ← what you actually pay   [D10B, used for yield calc]", False),
    ("", False),
    ("Approx Retail Ask Yield % (col G)", True),
    ("Annualised yield if you buy at the DMO dirty price (col I) and hold to maturity.", False),
    ("  Calculated via bisection at export time — static, does not update with overrides.", False),
    ("  Typically slightly lower than Effective Yield % because you pay a higher price.", False),
    ("  The gap between G and F represents the cost of using the DMO retail service.", False),
    ("", False),
    ("Cash flow columns (L–T)", True),
    ("L  Annual Coupon Cash (£)         = Nominal × Coupon % / 100", False),
    ("M  Capital Uplift to Par (£)       = Nominal × (100 − Price) / 100  [CGT-exempt]", False),
    ("N  Approx Gross Cash Gain (£)      = L × Years + M", False),
    ("O  Coupon Tax @20% (£)             = L × Years × 20%", False),
    ("P  Coupon Tax @40% (£)             = L × Years × 40%", False),
    ("Q  Coupon Tax @45% (£)             = L × Years × 45%", False),
    ("R  Approx Net Cash Gain @20% (£)   = N − O  [capital uplift is never taxed]", False),
    ("S  Approx Net Cash Gain @40% (£)   = N − P", False),
    ("T  Approx Net Cash Gain @45% (£)   = N − Q", False),
    ("", False),
    ("Approx Accrued Interest (col X)", True),
    ("The interest that has built up since the last coupon payment date.", False),
    ("Formula: (Nominal × Coupon% / 100 / 2) × (days since last coupon ÷ days in coupon period)", False),
    ("Coupon payment dates are derived from the maturity date — UK gilts pay on the same", False),
    ("  day and month as maturity, every 6 months (e.g. maturity Jan-31 → coupons Jan-31 and Jul-31).", False),
    ("Updates automatically every day via TODAY() — no need to regenerate the workbook.", False),
    ("This is an approximation using Actual/Actual day count.", False),
    ("Why it matters: when you buy a gilt, you pay the dirty price = clean price + accrued interest.", False),
    ("  The accrued interest is recovered when the next coupon is paid to you in full.", False),
    ("", False),
    ("Annual Net columns (U, V, W — shown in blue)", True),
    ("U  Annual Net @20% = R ÷ Years to Maturity", False),
    ("V  Annual Net @40% = S ÷ Years to Maturity", False),
    ("W  Annual Net @45% = T ÷ Years to Maturity", False),
    ("These are the primary comparison figures — after-tax cash per year, normalised for", False),
    ("  duration so gilts with different maturities are on a level playing field.", False),
    ("Without dividing by years, a 30-year gilt always looks better in total than a 2-year", False),
    ("  gilt, regardless of quality. Annual Net fixes this.", False),
    ("Note: Annual Net is a comparison metric — it is not the actual cash you receive each year.", False),
    ("  Actual annual cash = J (coupon only). Capital uplift arrives as a lump sum at maturity.", False),
    ("", False),
    ("How to rank gilts using the Analysis sheet", True),
    ("1. Click the autofilter arrow on column V (Annual Net @40%) and sort largest to smallest.", False),
    ("   (Use U for basic rate 20%, W for additional rate 45%.)", False),
    ("2. Use the Years to Maturity autofilter to restrict to your preferred holding window.", False),
    ("3. Among similarly-ranked gilts, prefer higher Capital Uplift (col K) — tax-free gain.", False),
    ("4. To rank by best pre-tax yield instead, sort column F (Effective Yield %) descending.", False),
    ("These columns update live when you change Nominal Amount or Override Price in Inputs.", False),
    ("", False),
    ("--- SETTINGS SHEET ---", True),
    ("", False),
    ("DefaultNominalAmount (row 2)", True),
    ("Change this single cell to set the default £ holding size for all gilts.", False),
    ("All gilts where Inputs col C is blank will use this value.", False),
    ("Tax scenario rates (rows 5–7) are reference only — Analysis uses hardcoded 20/40/45%.", False),
    ("", False),
    ("--- SELLING BEFORE MATURITY ---", True),
    ("", False),
    ("You do not have to hold a gilt to maturity. If you sell early:", False),
    ("  - All coupons already paid are yours to keep.", False),
    ("  - Accrued interest since the last coupon is paid by the buyer automatically.", False),
    ("  - Your capital outcome depends on the price at the date you sell.", False),
    ("    If rates have risen, the price will be lower — potential capital loss.", False),
    ("    If rates have fallen, the price will be higher — capital gain (still CGT-exempt).", False),
    ("To model an early exit: enter your expected sale price in Override Price (Inputs col D).", False),
    ("  The Analysis sheet will recalculate using that price instead of £100 par.", False),
)


def _populate_instructions(sheet: Worksheet) -> None:
    sheet.column_dimensions["A"].width = 90
    for text, is_heading in _INSTRUCTIONS:
        sheet.append((text,))
        if text and is_heading:
            sheet.cell(sheet.max_row, 1).font = Font(bold=True, size=11)
        elif text:
            sheet.cell(sheet.max_row, 1).alignment = Alignment(wrap_text=True)
    sheet.freeze_panes = None  # no freeze on instructions
    sheet.auto_filter.ref = None
